# -*- coding: utf-8 -*-
"""Assignment01.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1W7zo10oUzGBNIcGsCqLCFxG1VvUHBZfE
"""

###Revision-01-Login and registration system

x=[]
spechar = ["%", "!", "@"]
q = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
class Login:

    def __init__(self,name,password):
        self.name=name
        self.password=password

    # Validating Username:

    def user(self):
        if self.name.endswith("@gmail.com" or "@yahoo.in" or "rediffmail.com" or "hotmail.com"):
            x.append(self.name)
            return self.name

    # finding one special character in password:

    def special(self):
        for i in spechar:
            b = self.password.find(i)
            if b >= 0:
                return self.password


    # finding one upper character in password:
    def upper(self):
        for i in self.password:
            y = i.isupper()
            if y == True:
                return self.password

    # finding one lower character in password:
    def lower(self):
        for i in self.password:
            z = i.islower()
            if z == True:
                return self.password

    # finding one number in string in password:


    def onenumber(self):
        for i in q:
            b = self.password.find(i)
            if b >= 0:
                return self.password

n=int(input("-----------------------------\nD27 BATCH LOGIN AND REGISTRATION SYSTEM\n1.  Login\n2.  Registration\n3.  Forget Password\n\n Enter  "
        "index of the operation to perform.......1 or 2 or 3....:"))
print("\n\n\n")
if n==1:
    while True:
        f = input("Enter valid email/username......:")
        g = input("Enter password..................:")
        person1 = Login(f, g)
        temp=person1.user()
        pass1=person1.special()
        pass2=person1.upper()
        pass3=person1.lower()
        pass4=person1.onenumber()
        if temp==f and (pass1 and pass2 and pass3 and pass4)==g and (len(g) in range(6,17)):
            print(".........\nLOGIN SUCCESSFULL\n.......!!")
            break
        else:
            print(".........\nLOGIN FAILED...xxxx\n\nEnter Valid Credentials!!\n.......!!")
elif n==2:
    while True:
        f=input("Enter valid email/username......:")
        g=input("Enter password..................:")
        person1 = Login(f, g)
        temp=person1.user()
        pass1=person1.special()
        pass2=person1.upper()
        pass3=person1.lower()
        pass4=person1.onenumber()
        if temp==f and (pass1 and pass2 and pass3 and pass4)==g and (len(g) in range(6,17)):
            print(".........\nREGISTRATION SUCCESSFULL\n.......!!")
            break
        else:
            print(".........\nREGISTRATION UNSUCCESSFULL\n\nEnter Valid Credentials!!\n.......!!")

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("NewGrades.xlsx")

###Revision-02-Login and registration system dated 15-03-2022

x=[]
spechar = ["%", "!", "@"]
q = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
class Login:

    def __init__(self,name,password):
        self.name=name
        self.password=password

    # Validating Username:

    def user(self):
        if self.name.endswith("@gmail.com" or "@yahoo.in" or "rediffmail.com" or "hotmail.com"):
            x.append(self.name)
            return self.name

    # finding one special character in password:

    def special(self):
        for i in spechar:
            b = self.password.find(i)
            if b >= 0:
                return self.password


    # finding one upper character in password:
    def upper(self):
        for i in self.password:
            y = i.isupper()
            if y == True:
                return self.password

    # finding one lower character in password:
    def lower(self):
        for i in self.password:
            z = i.islower()
            if z == True:
                return self.password

    # finding one number in string in password:


    def onenumber(self):
        for i in q:
            b = self.password.find(i)
            if b >= 0:
                return self.password

n=int(input("-----------------------------\nD27 BATCH LOGIN AND REGISTRATION SYSTEM\n1.  Login\n2.  Registration\n3.  Forget Password\n\n Enter  "
        "index of the operation to perform.......1 or 2 or 3....:"))
print("\n\n\n")
if n==1:
    while True:
        f = input("Enter valid email/username......:")
        import os
        if os.path.exists(f+".txt"):
            g = input("Enter password..................:")
            dummypass=open(f+".txt","r")
            passcode=dummypass.read()
            if passcode==g:
              print(".........\nLOGIN SUCCESSFULL\n.......!!")
            else:
              print("Password incorrect")
            break
        else:
            print(".........\n User Name not found..Kindly please Register!!\n.......!!")
            break
elif n==2:
    while True:
        f=input("Enter valid email/username......:")
        g=input("Enter password..................:")
        person1 = Login(f, g)
        temp=person1.user()
        pass1=person1.special()
        pass2=person1.upper()
        pass3=person1.lower()
        pass4=person1.onenumber()
        if temp==f and (pass1 and pass2 and pass3 and pass4)==g and (len(g) in range(6,17)):
            print(".........\nREGISTRATION SUCCESSFULL\n.......!!")
            dummy=open(f+".txt","w")
            dummy.write(g)
            dummy.close()
            break
        else:
            print(".........\nREGISTRATION UNSUCCESSFULL\n\nEnter Valid Credentials!!\n.......!!")



###Revision-03-Login and registration system dated 15-03-2022

x=[]
spechar = ["%", "!", "@"]
q = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
class Login:

    def __init__(self,name,password):
        self.name=name
        self.password=password

    # Validating Username:

    def user(self):
        if self.name.endswith("@gmail.com" or "@yahoo.in" or "rediffmail.com" or "hotmail.com"):
            x.append(self.name)
            return self.name

    # finding one special character in password:

    def special(self):
        for i in spechar:
            b = self.password.find(i)
            if b >= 0:
                return self.password


    # finding one upper character in password:
    def upper(self):
        for i in self.password:
            y = i.isupper()
            if y == True:
                return self.password

    # finding one lower character in password:
    def lower(self):
        for i in self.password:
            z = i.islower()
            if z == True:
                return self.password

    # finding one number in string in password:


    def onenumber(self):
        for i in q:
            b = self.password.find(i)
            if b >= 0:
                return self.password
while True:
    n=int(input("-----------------------------\nD27 BATCH LOGIN AND REGISTRATION SYSTEM\n1.  Login\n2.  Registration\n3.  Forget Password\n 4. To exit the program or backspace, Enter-'0' \n Enter  "
            "index of the operation to perform.......1 or 2 or 3 or 0...:"))
    print("\n\n\n")
    if n==1:
        while True:
            f = input("Enter valid email/username......:")
            import os
            if os.path.exists(f+".txt"):
                g = input("Enter password..................:")
                dummypass=open(f+".txt","r")
                passcode=dummypass.read()
                if passcode==g:
                  print(".........\nLOGIN SUCCESSFULL\n.......!!")
                else:
                  print("Password incorrect")
                break
            else:
                print(".........\n User Name not found..Kindly please Register!!\n.......!!")
               
    elif n==2 :
        while True:
            f=input("Enter valid email/username......:")
            g=input("Enter password..................:")
            person1 = Login(f, g)
            temp=person1.user()
            pass1=person1.special()
            pass2=person1.upper()
            pass3=person1.lower()
            pass4=person1.onenumber()
            if temp==f and (pass1 and pass2 and pass3 and pass4)==g and (len(g) in range(6,17)):
                print(".........\nREGISTRATION SUCCESSFULL\n.......!!")
                dummy=open(f+".txt","w")
                dummy.write(g)
                dummy.close()
                break
            else:
                print(".........\nREGISTRATION UNSUCCESSFULL\n\nEnter Valid Credentials!!\n.......!!")
    elif n==3:
          while True:
            f = input("Enter valid email/username to retrive Password......:")
            import os
            if os.path.exists(f+".txt"):
                dummypass=open(f+".txt","r")
                print("-------\nThe Previous password\n---------")
                print(dummypass.read())
                break
            print("Enter Valid User name/email id present in the system...........!!!")
    else:         
       break